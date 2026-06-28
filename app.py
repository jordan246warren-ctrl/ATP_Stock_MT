"""ATP Inventory Engine — Streamlit frontend."""

import io
import re
from datetime import date
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from src.orders import ORDER_DISPLAY_COLS, ORDER_TIER_COLS
from src.parser import AGE_BUCKETS, NUMERIC_COLS, format_cell, grade_family_key, process_inventory_views
from src.old_stock import OldStockItem, extract_old_stock


# ── Display helpers ───────────────────────────────────────────────────────────

def _read_app_version() -> str:
    version_path = Path(__file__).resolve().parent / "VERSION"
    if version_path.is_file():
        return version_path.read_text(encoding="utf-8").strip()
    return "dev"


def _build_display_df(
    df: pd.DataFrame,
    rollups: dict[str, dict[str, int]],
) -> pd.DataFrame:
    """Apply bracket formatting: direct (total) where rollups exist."""
    rows = []
    for _, row in df.iterrows():
        group = row["Selling Group"]
        incoming = rollups.get(group, {})
        formatted = {"Selling Group": group}
        for col in NUMERIC_COLS:
            direct = int(row[col])
            formatted[col] = format_cell(direct, incoming.get(col, 0))
        rows.append(formatted)
    return pd.DataFrame(rows, columns=["Selling Group", *NUMERIC_COLS])


def _is_spacer_row(group: str) -> bool:
    return group == ""


def _is_total_row(group: str) -> bool:
    return group == "TOTAL"


def _append_footer_total(
    df_display: pd.DataFrame,
    df_raw: pd.DataFrame,
    numeric_cols: list[str],
) -> pd.DataFrame:
    """Append a TOTAL row summing numeric columns from the raw (pre-format) data."""
    total_row: dict = {"Selling Group": "TOTAL"}
    for col in numeric_cols:
        total_row[col] = int(df_raw[col].sum()) if col in df_raw.columns else 0
    return pd.concat([df_display, pd.DataFrame([total_row])], ignore_index=True)


def _needs_spacer(prev: str | None, curr: str) -> bool:
    if prev is None:
        return False
    if grade_family_key(curr) != grade_family_key(prev):
        return True
    if curr.startswith("Bull ") and not prev.startswith("Bull "):
        return True
    return False


def _spacer_row(col_names: list[str]) -> dict:
    return {"Selling Group": "", **{col: "" for col in col_names}}


def _insert_group_spacers(df: pd.DataFrame, numeric_cols: list[str]) -> pd.DataFrame:
    """Insert blank rows between grade families and before the Bull block."""
    rows: list[dict] = []
    prev_group: str | None = None
    all_cols = ["Selling Group", *numeric_cols]
    for _, row in df.iterrows():
        group = row["Selling Group"]
        if _needs_spacer(prev_group, group):
            rows.append(_spacer_row(numeric_cols))
        rows.append(row.to_dict())
        prev_group = group
    return pd.DataFrame(rows, columns=all_cols)


_SPACER_STYLE = (
    "background-color: transparent; border: none; "
    "padding: 4px 14px; text-align: center;"
)

_TOTAL_ROW_STYLE = (
    "background-color: #ffffff; color: #000000; "
    "font-weight: bold; text-align: center;"
)


def _parse_cell_direct(val) -> int | None:
    """Extract the leading direct count from a formatted cell value."""
    if isinstance(val, str) and val and val != "—":
        m = re.match(r"(-?\d+)", val)
        if m:
            return int(m.group(1))
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def _cell_style(val) -> str:
    if val == "—":
        return (
            "background-color: transparent; color: #4b4b52; "
            "text-align: center;"
        )
    direct = _parse_cell_direct(val)
    if direct is not None:
        if direct == 0:
            return (
                "background-color: transparent; color: #4b4b52; "
                "text-align: center;"
            )
        if direct > 0:
            return (
                "background-color: #ffffff; color: #000000; "
                "font-weight: bold; text-align: center;"
            )
        return (
            "background-color: #ffffff; color: #ff0000; "
            "font-weight: bold; text-align: center;"
        )
    return ""


def _style_numeric_row(df: pd.DataFrame):
    def style_row(row: pd.Series) -> list[str]:
        group = df.at[row.name, "Selling Group"]
        if _is_spacer_row(group):
            return [_SPACER_STYLE] * len(row)
        if _is_total_row(group):
            return [_TOTAL_ROW_STYLE] * len(row)
        return [_cell_style(v) for v in row]

    return style_row


def _inject_spacer_row_classes(html: str, df: pd.DataFrame) -> str:
    """Add grade-spacer class to blank separator rows in styler HTML."""
    spacer_positions = {
        i + 1
        for i, row in enumerate(df.itertuples(index=False))
        if row[0] == ""
    }
    if not spacer_positions:
        return html

    tbody_pat = re.compile(r"(<tbody>)(.*?)(</tbody>)", re.DOTALL)

    def replacer(match: re.Match[str]) -> str:
        trs = re.findall(r"<tr[^>]*>.*?</tr>", match.group(2), re.DOTALL)
        new_trs = []
        for i, tr in enumerate(trs, 1):
            if i in spacer_positions:
                tr = re.sub(r"^<tr\b", '<tr class="grade-spacer"', tr, count=1)
            new_trs.append(tr)
        return match.group(1) + "".join(new_trs) + match.group(3)

    return tbody_pat.sub(replacer, html)


def _render_styled_table(df_display: pd.DataFrame, numeric_cols: list[str]) -> None:
    """Render a styled HTML table inside the atp-table wrapper."""
    styler = df_display.style
    try:
        styler = styler.apply(
            _style_numeric_row(df_display),
            axis=1,
            subset=numeric_cols,
        )
    except AttributeError:
        styler = styler.applymap(_cell_style, subset=numeric_cols)

    spacer_idx = df_display.index[df_display["Selling Group"] == ""].tolist()
    styler = styler.set_properties(subset=["Selling Group"], **{"text-align": "left"})
    if spacer_idx:
        styler = styler.set_properties(
            subset=pd.IndexSlice[spacer_idx, :],
            **{
                "background-color": "transparent",
                "border": "none",
                "padding": "4px 14px",
            },
        )
    styler = styler.hide(axis="index")

    table_html = _inject_spacer_row_classes(styler.to_html(), df_display)
    st.markdown(f'<div class="atp-table">{table_html}</div>', unsafe_allow_html=True)


def _render_inventory_tab(
    df: pd.DataFrame,
    rollups: dict[str, dict[str, int]],
    numeric_cols: list[str],
    *,
    empty_message: str,
    download_label: str,
    footer_total: bool = False,
) -> None:
    """Render a bracket-formatted inventory table with optional Excel download."""
    if df.empty:
        st.warning(empty_message)
        return

    df_body_display = _build_display_df(df, rollups)
    if "Total Physical Stock" not in numeric_cols:
        df_body_display = df_body_display.drop(columns=["Total Physical Stock"])
    df_display = _insert_group_spacers(df_body_display, numeric_cols)
    if footer_total:
        df_display = _append_footer_total(df_display, df, numeric_cols)
    _render_styled_table(df_display, numeric_cols)

    xlsx_bytes = _build_excel(df_display)
    st.download_button(
        label="⬇ Download Excel Report",
        data=xlsx_bytes,
        file_name=f"ATP_{download_label.replace(' ', '_')}_{date.today().strftime('%d-%m-%y')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _render_orders_tab(orders_df: pd.DataFrame) -> None:
    """Render the tier-demand Orders table with optional Excel download."""
    if orders_df.empty:
        st.warning("No mapped order lines found in the uploaded orders file.")
        return

    df_display = _insert_group_spacers(orders_df, ORDER_DISPLAY_COLS)
    _render_styled_table(df_display, ORDER_DISPLAY_COLS)

    xlsx_bytes = _build_excel(df_display)
    st.download_button(
        label="⬇ Download Excel Report",
        data=xlsx_bytes,
        file_name=f"ATP_Orders_Report_{date.today().strftime('%d-%m-%y')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Excel export helper ───────────────────────────────────────────────────────

def _build_excel(df_display: pd.DataFrame) -> bytes:
    """Build a formatted .xlsx workbook from df_display."""
    buf = io.BytesIO()
    df_display.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb.active

    thin = Side(style="thin")
    all_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    col_count = len(df_display.columns)
    if col_count == 6:
        col_widths = {1: 26, 2: 22, 3: 12, 4: 12, 5: 12, 6: 12}
    else:
        col_widths = {1: 26, 2: 12, 3: 12, 4: 12, 5: 12}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width

    for row in ws.iter_rows():
        row_num = row[0].row
        is_header = row_num == 1
        is_spacer = not is_header and (row[0].value in (None, ""))

        ws.row_dimensions[row_num].height = 8 if is_spacer else 20

        for cell in row:
            if is_spacer:
                cell.border = Border()
                continue

            cell.border = all_thin

            if is_header:
                cell.alignment = center
                cell.font = Font(bold=True)
                continue

            if cell.column == 1:
                cell.alignment = left
            else:
                cell.alignment = center
                if cell.value == "—":
                    cell.font = Font(color="888888")
                elif cell.value not in (None, ""):
                    direct = _parse_cell_direct(cell.value)
                    if direct is not None and direct < 0:
                        cell.font = Font(bold=True, color="FF0000")
                    else:
                        cell.font = Font(bold=True)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def _build_old_stock_excel(items: list[OldStockItem]) -> bytes:
    """Build a formatted .xlsx workbook from old-stock items (6-9 and 9+ buckets)."""
    rows = [
        {
            "Selling Group": it.selling_group,
            "Days Old": round(it.days_old, 2),
            "Barcode": it.barcode,
            "Bucket": it.bucket,
        }
        for it in items
    ]
    df = pd.DataFrame(rows, columns=["Selling Group", "Days Old", "Barcode", "Bucket"])

    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb.active

    thin = Side(style="thin")
    all_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    col_widths = {1: 26, 2: 12, 3: 20, 4: 10}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width

    for row in ws.iter_rows():
        row_num = row[0].row
        ws.row_dimensions[row_num].height = 18
        for cell in row:
            cell.border = all_thin
            if row_num == 1:
                cell.alignment = center
                cell.font = Font(bold=True)
            elif cell.column == 1:
                cell.alignment = left
            else:
                cell.alignment = center

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ── Old-stock dialogs ─────────────────────────────────────────────────────────

@st.dialog("6-9 Day Aged Stock")
def _show_6_9_dialog(items: list[OldStockItem]) -> None:
    filtered = [it for it in items if it.bucket == "6-9"]
    if not filtered:
        st.info("No items in the 6-9 day bucket.")
        return
    st.caption(f"{len(filtered)} item(s) — sorted oldest first")
    lines = "\n".join(
        f"{it.selling_group};  {it.days_old:.2f} days;  {it.barcode}"
        for it in filtered
    )
    st.code(lines, language=None)


@st.dialog("9+ Day Aged Stock")
def _show_9plus_dialog(items: list[OldStockItem]) -> None:
    filtered = [it for it in items if it.bucket == "9+"]
    if not filtered:
        st.info("No items in the 9+ day bucket.")
        return
    st.caption(f"{len(filtered)} item(s) — sorted oldest first")
    lines = "\n".join(
        f"{it.selling_group};  {it.days_old:.2f} days;  {it.barcode}"
        for it in filtered
    )
    st.code(lines, language=None)


# ── Streamlit app ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="ATP Inventory Engine",
    page_icon="📦",
    layout="wide",
)

st.markdown(
    """
    <style>
        /* Screen: hide Streamlit chrome */
        #MainMenu, header, footer { visibility: hidden; }

        /* ── Screen: ATP table ───────────────────────────────────────────── */
        .atp-table table {
            border-collapse: collapse;
            width: 100%;
            font-family: Arial, sans-serif;
            font-size: 14px;
        }
        .atp-table th {
            padding: 10px 14px;
            text-align: center;
            background-color: #0e1117;
            color: #fafafa;
            border-bottom: 2px solid #444444;
        }
        .atp-table th:first-child { text-align: left; }
        .atp-table td {
            padding: 8px 14px;
            text-align: center;
            border-bottom: 1px solid #2d2d3a;
        }
        .atp-table td:first-child { text-align: left; }

        /* Grade-family spacer rows */
        .atp-table tr.grade-spacer td {
            border-bottom: none !important;
            padding: 4px 14px !important;
            height: 10px;
            background-color: transparent !important;
        }

        /* ── Print: A4 portrait ──────────────────────────────────────────── */
        @media print {
            @page { size: A4 portrait; margin: 1cm; }

            header, footer, #MainMenu { display: none !important; }

            [data-testid="stHeading"],
            [data-testid="stCaption"],
            [data-testid="stFileUploader"],
            [data-testid="stDownloadButton"],
            [data-testid="stButton"],
            [data-testid="stHorizontalBlock"]:has([data-testid="stButton"]),
            [data-testid="stAlert"],
            [data-testid="stSpinner"],
            hr {
                display: none !important;
            }

            .stApp, [data-testid="stAppViewContainer"],
            [data-testid="stMain"], section.main {
                overflow: visible !important;
            }

            .atp-table {
                position: static !important;
                width: 100% !important;
            }

            .atp-table table {
                border-collapse: collapse !important;
                width: 100% !important;
                font-size: 9.5pt !important;
                font-family: Arial, sans-serif !important;
            }

            .atp-table th,
            .atp-table td {
                border: 1px solid black !important;
                background-color: white !important;
                color: black !important;
                padding: 6px 10px !important;
                text-align: center !important;
                line-height: 1.5 !important;
            }

            .atp-table th { font-weight: bold !important; }
            .atp-table td:first-child { text-align: left !important; }

            .atp-table thead { display: table-header-group !important; }
            .atp-table tr { break-inside: avoid !important; }

            .atp-table tr.grade-spacer td {
                border: none !important;
                padding: 4px 10px !important;
                height: 10px !important;
                break-inside: avoid !important;
            }
        }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("ATP Inventory Engine")
st.caption(
    f"Available to Promise — physical stock by selling group and age bucket · v{_read_app_version()}"
)

soh_col, orders_col = st.columns(2)
with soh_col:
    uploaded_soh = st.file_uploader(
        "Upload Stock on Hand (SOH)",
        type=["xls", "xlsx"],
        help="Drag and drop your scanned SOH file here (.xls or .xlsx).",
    )
with orders_col:
    uploaded_orders = st.file_uploader(
        "Upload Abaserve Orders (optional)",
        type=["xls", "xlsx"],
        help="Upload to subtract open orders and display Available to Promise (ATP).",
    )

st.divider()

if uploaded_soh is None and uploaded_orders is None:
    st.info("Upload an SOH and/or Orders file above.")
else:
    try:
        with st.spinner("Processing…"):
            views = process_inventory_views(uploaded_soh, uploaded_orders)
            old_items = []
            if uploaded_soh is not None:
                uploaded_soh.seek(0)
                old_items = extract_old_stock(uploaded_soh)

        tab_labels: list[str] = []
        if uploaded_soh is not None:
            tab_labels.append("Physical Stock")
        if uploaded_soh is not None and uploaded_orders is not None:
            tab_labels.append("ATP")
        if uploaded_orders is not None:
            tab_labels.append("Orders")

        tabs = st.tabs(tab_labels)
        tab_map = dict(zip(tab_labels, tabs))

        if "Physical Stock" in tab_map:
            with tab_map["Physical Stock"]:
                st.subheader("Physical Stock")
                _render_inventory_tab(
                    views.physical_df,
                    views.physical_rollups,
                    ["Total Physical Stock", *AGE_BUCKETS],
                    empty_message=(
                        "No stock found after processing. "
                        "Verify that Groups.csv contains the product codes in this SOH file."
                    ),
                    download_label="Stock Report",
                    footer_total=True,
                )
                if old_items:
                    n_6_9 = sum(1 for it in old_items if it.bucket == "6-9")
                    n_9plus = sum(1 for it in old_items if it.bucket == "9+")
                    view_col_69, view_col_9plus, old_dl_col = st.columns([1, 1, 1])
                    with view_col_69:
                        if n_6_9 and st.button(f"View 6-9 Aged Stock ({n_6_9})"):
                            _show_6_9_dialog(old_items)
                    with view_col_9plus:
                        if n_9plus and st.button(f"View 9+ Aged Stock ({n_9plus})"):
                            _show_9plus_dialog(old_items)
                    with old_dl_col:
                        old_xlsx_bytes = _build_old_stock_excel(old_items)
                        st.download_button(
                            label="⬇ Download Old Stock Report",
                            data=old_xlsx_bytes,
                            file_name=f"Old_Stock_Report_{date.today().strftime('%d-%m-%y')}.xlsx",
                            mime=(
                                "application/vnd.openxmlformats-officedocument"
                                ".spreadsheetml.sheet"
                            ),
                        )

        if "ATP" in tab_map:
            with tab_map["ATP"]:
                st.subheader("Available to Promise")
                _render_inventory_tab(
                    views.atp_df,
                    views.atp_rollups,
                    list(AGE_BUCKETS),
                    empty_message="No ATP rows to display after applying order deductions.",
                    download_label="ATP Report",
                )

        if "Orders" in tab_map:
            with tab_map["Orders"]:
                st.subheader("Orders by Customer Tier")
                _render_orders_tab(views.orders_df)

        if views.unmapped_soh:
            codes = ", ".join(f"`{c}`" for c in views.unmapped_soh)
            st.warning(
                f"**{len(views.unmapped_soh)} unmapped SOH product code(s)** were skipped — "
                f"they appear in the SOH but are missing from Groups.csv. "
                f"Add them to Groups.csv and re-upload to include them.\n\n"
                f"{codes}"
            )

        if views.unmapped_orders:
            codes = ", ".join(f"`{c}`" for c in views.unmapped_orders)
            st.warning(
                f"**{len(views.unmapped_orders)} unmapped order product code(s)** were skipped — "
                f"they appear in the orders file but are missing from Groups.csv.\n\n"
                f"{codes}"
            )

        if views.missing_rollup_targets:
            issues = "\n".join(f"- {m}" for m in views.missing_rollup_targets)
            st.warning(
                f"**{len(views.missing_rollup_targets)} rollup mapping issue(s)** — "
                f"some grade families are missing required target groups in Groups.csv:\n\n"
                f"{issues}"
            )

    except ValueError as exc:
        st.error(f"Could not parse file: {exc}")
    except Exception as exc:
        st.error(f"Unexpected error: {exc}")
